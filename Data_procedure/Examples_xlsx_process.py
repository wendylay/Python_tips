"""
with openpyxl
"""
import openpyxl


def write_excel_xlsx(path, sheet_name, value):
    workbook = openpyxl.Workbook()
    sheet = workbook.active  # 获取当前的工作表, 在新建的表中就不会出现第一个sheet没有使用的情况
    sheet.title = sheet_name
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            # Row or column values must be at least 1, the box in excel called cell (same as matlab)
            sheet.cell(row=i + 1, column=j + 1, value=value[i][j])
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def addExcel(path, value, sheet):
    """
    :param path: 写入excel的路径
    :param value: 追加的数据
    :param sheet: sheet的名称
    :return:
    """

    wb = openpyxl.load_workbook(path)
    wb.create_sheet(sheet)
    ws = wb[sheet]

    for ss in value:
        ws.append(ss)  # append list value
    wb.save(path)
    print("写入成功")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


book_name_xlsx = 'xlsx格式测试工作簿.xlsx'
sheet_name_xlsx = 'xlsx格式测试表'
value3 = [["姓名", "性别", "年龄", "城市", "职业"],
          ["111", "女", "66", "石家庄", "运维工程师"],
          ["222", "男", "55", "南京", "饭店老板"],
          ["333", "女", "27", "苏州", "保安"], ]

write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
addExcel(book_name_xlsx, value3, sheet_name_xlsx + '_append')

"""
with pandas
"""

# method 1: write at the same time
# refer: https://blog.csdn.net/weixin_43060843/article/details/100766677?utm_medium=distribute.pc_relevant.none-task-blog-BlogCommendFromBaidu-3.control&depth_1-utm_source=distribute.pc_relevant.none-task-blog-BlogCommendFromBaidu-3.control
import pandas as pd
import os

df1 = pd.DataFrame([1, 2, 3, 4])
df2 = pd.DataFrame([1, 2, 3, 4])
df3 = pd.DataFrame([1, 2, 3, 4])
writer = pd.ExcelWriter('test.xlsx')
df1.to_excel(writer, "test1")
df2.to_excel(writer, "test2")
df3.to_excel(writer, "test3")
writer.save()

# method 2: append the sheet to exist workbook
# refer: https://blog.csdn.net/unsterbliche/article/details/80578606
import pandas as pd
from openpyxl import load_workbook

df1 = pd.DataFrame([1, 2, 3, 4])
excelWriter = pd.ExcelWriter('test.xlsx',
                             engine='openpyxl')
df1.to_excel('test.xlsx', index=False)


def excelAddSheet(dataframe, excelWriter, sheet_name):
    book = load_workbook(excelWriter.path)
    excelWriter.book = book  # the way to avoid overwrite
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sheet_name, index=None)
    excelWriter.close()


# excel必需已经存在，因此先建立一个空的sheet
excelAddSheet(df1, excelWriter, 'test2')
